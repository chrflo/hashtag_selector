from openpyxl import load_workbook
import random
import datetime
import sys
import logger.logging as logger
import arguments.argumentParser as parser

# Constants
HEADER_ROW = 1
HASHTAGS_IN_COL_ROW = 2
HASHTAGS_TO_USE_ROW = 3
START_ROW = 4
CATEGORY_NUM = 9
COL_NUM = CATEGORY_NUM * 2
END_COL = COL_NUM + 1
START_COL = 2
STEP = 2
SHEET_NAME = 'SheetWeighted'
SPACER = '•'
FILENAME = 'hashtags'
FILE_EXT = 'txt'


def normalize(weights):
    assert len(weights) > 0

    total = sum(weights)

    tempWeight = []
    for i in range(0, len(weights)):
        tempWeight.append(weights[i]/total)

    return tempWeight


def weighted_choice(seq, weights):
    log.debug('Seq Length: {1}, Weight Length: {0}, Probability Dist: {2}'.format(len(weights), len(seq), 1. - sum(weights)))
    assert len(weights) == len(seq)
    assert abs(1. - sum(weights)) < 1e-6

    x = random.random()
    for i, elmt in enumerate(seq):
        if x <= weights[i]:
            return elmt
        x -= weights[i]


def get_hashtags(count, weighted=False, filename=FILENAME):
    assert count != 0

    # open the file to write to and empty it's current contents
    file = open('{0}_{1}.{2}'.format(filename, datetime.date.today(), FILE_EXT), 'a')
    file.truncate(0)


    for m in range(0, count):
        hashtags = []
        rejected = []
        for i in range(START_COL, END_COL, STEP):
            r = ws.cell(row=HASHTAGS_IN_COL_ROW, column=i).value
            num = ws.cell(row=HASHTAGS_TO_USE_ROW, column=i).value

            tempHashtags = []
            tempWeight = []
            log.debug('Number to pick: {0}, Rows: {1}'.format(num, r))
            for j in range(START_ROW, START_ROW+r):
                tempHashtags.append(ws.cell(row=j, column=i).value)
                tempWeight.append(ws.cell(row=j, column=i+1).value)

            #randomly select the hashtag
            # for k in range(0, num):
            k = 0
            while k < num:
                k += 1
                if not weighted:
                    index = random.randrange(len(tempHashtags))
                    tag = '#' + tempHashtags.pop(index)

                    if hashtagChecker(tag):
                        hashtags.append(tag)
                    else:
                        rejected.append(tag)
                        k -= 1
                else:
                    # normalize the weights
                    # this is placed here to ensure that the weights are normalized
                    # prior to entering the random weighted selection
                    tempWeight = normalize(tempWeight)

                    weightedTag = weighted_choice(tempHashtags, tempWeight)
                    # let's remove the weighted tag choice from the temp hashtags
                    tagIndex = tempHashtags.index(weightedTag)
                    tempHashtags.pop(tagIndex)
                    tempWeight.pop(tagIndex)

                    tag = '#' + weightedTag
                    if hashtagChecker(tag):
                        hashtags.append(tag)
                    else:
                        rejected.append(tag)
                        k -= 1

        # update the set and overview
        updateHashtagSet(hashtags)

        hashtag = '\n{1}\n{0}\n{0}\n{0}\n'.format(SPACER, len(hashtags)) + ' '.join(map(str, hashtags))
        print('Tags: {1}, Rejected: {0}\n'.format(len(rejected), len(hashtags)) + ' '.join(map(str, rejected)))
        file.write(hashtag)

    file.close()


# Using the same hashtags more than 3 times within the same 24 hours - quantify 24 hours (2 posts)
# Using the same exact hashtag sets more than 3 times in a row
# Using any hashtag more than 6 times in your last 10 posts
def hashtagChecker (tag):
    if tag not in hashtagoverview or len(hashtagsets) <= 2:
        return True

    inverse = hashtagsets[::-1]
    twopost = False
    for i in range(0, 2):
        if tag in inverse[i]:
            twopost = True
            break

    return (hashtagoverview[tag] < 6) and (not twopost)

# TODO: this is not the most efficient way, think of better way later
def updateHashtagSet (tags):
    if len(hashtagsets) >= 11:
        tmpTags = hashtagsets.pop(0)
        for i in range(0, len(tmpTags)):
            tag = tmpTags[i] #[1:]
            assert tag in hashtagoverview

            # decrement the count and if 0 remove
            if hashtagoverview.get(tag) == 1:
                del hashtagoverview[tag]
            else:
                hashtagoverview[tag] -= 1

    hashtagsets.append(tags)

    # update the tags dict
    for i in range(0, len(tags)):
        tag = tags[i] #[1:]
        # increment the count and if 0 remove
        if tag in hashtagoverview:
            hashtagoverview[tag] += 1
        else:
            hashtagoverview[tag] = 1


appName = sys.argv[0]
args = parser.ArgsParser().parse(sys.argv[1:])

# init logging

global log
log = logger.Logger(args.logLvl, args.logName, args.logPath).logger

wb = load_workbook(args.workbook, data_only=True)
ws = wb[SHEET_NAME]
hashtagsets = []
hashtagoverview = {}
get_hashtags(args.iterations, args.weight, args.filename)

